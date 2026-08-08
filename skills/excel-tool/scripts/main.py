#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 工具 — 读取/写入 .xlsx 文件。"""
import argparse
import json
import sys

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Excel 工具")
    parser.add_argument("--action", required=True, choices=["read", "write"], help="操作")
    parser.add_argument("--file", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet", default="", help="工作表名称")
    parser.add_argument("--data", default="", help="JSON 数据（write 时必填）")
    args = parser.parse_args()

    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        print(json.dumps({"error": "需要 openpyxl 库: pip install openpyxl"}, ensure_ascii=False))
        sys.exit(1)

    if args.action == "read":
        try:
            wb = load_workbook(args.file, data_only=True)
            ws = wb[args.sheet] if args.sheet else wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            headers = rows[0] if rows else []
            data = [dict(zip(headers, r)) for r in rows[1:]]
            print(json.dumps({"sheet": ws.title, "headers": headers, "rows": data, "count": len(data)}, ensure_ascii=False, indent=2))
        except FileNotFoundError:
            print(json.dumps({"error": f"文件不存在: {args.file}"}, ensure_ascii=False))
        except Exception as e:
            print(json.dumps({"error": str(e)}, ensure_ascii=False))

    elif args.action == "write":
        try:
            data = json.loads(args.data) if isinstance(args.data, str) else args.data
            if not isinstance(data, list):
                data = [data]
            wb = Workbook()
            ws = wb.active
            ws.title = args.sheet or "Sheet1"
            # 写表头
            headers = list(data[0].keys()) if data else []
            ws.append(headers)
            # 写数据
            for row in data:
                ws.append([row.get(h, "") for h in headers])
            wb.save(args.file)
            print(json.dumps({"success": True, "rows_written": len(data), "output_file": args.file}, ensure_ascii=False))
        except Exception as e:
            print(json.dumps({"error": str(e)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
